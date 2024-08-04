from openpyxl import load_workbook
import os

terms = './raw/structure/terms.txt'
excels = './raw/english/'


def read_terms():
    with open(terms, 'r') as reader:
        lines = reader.readlines()
    lines = [line.lower().strip() for line in lines]
    return lines


def convert(term):
    return f'''<span class="idef" data-idef="{term.capitalize()}">{term.capitalize()}</span> '''


def read_excels(terms):
    os.makedirs('./changed/', exist_ok=True)
    files = os.listdir(excels)
    for ifile in files:
        print('------------ File:: {} ---------------'.format(ifile))
        path = os.path.join(excels, ifile)
        wb = load_workbook(filename=path)
        info = {
            'Information cards': 'Statement',
            'Timeline': 'Description',
        }
        positions = {
            'Statement': 1,
            'Description': 2,
        }
        sheets = wb.sheetnames
        for sheet in sheets:
            print('xxxxxxx  Reading Sheet:: {} xxxxxxxxxx'.format(sheet))
            sheet_data = wb[sheet]  # the particular sheet
            name = info[sheet]  # get the column for the sheet
            col_id = positions[name]  # get the name for the column
            for row_cells in sheet_data.iter_rows(min_row=1, max_row=sheet_data.max_row):
                data = row_cells[col_id].value
                if data is not None:
                    values = data.split()
                    nvalues = []
                    for value in values:
                        if value in terms:
                            print(value)
                            value = convert(value)
                        nvalues.append(value)
                    row_cells[col_id].value = ' '.join(nvalues)
        save_file = os.path.join('./changed/english/', ifile)
        wb.save(save_file)


def main():
    read_excels(read_terms())


if __name__ == '__main__':
    main()
